"""XLSX file generation for check results."""

from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils.dataframe import dataframe_to_rows

from app.config import get_settings
from app.services.instagram_scraper import InstagramUser
from app.utils.logger import logger

settings = get_settings()


def create_styled_workbook(
    target_username: str,
    followers: list[InstagramUser],
    following: list[InstagramUser],
    non_mutual: list[InstagramUser],
) -> Workbook:
    """Create a styled Excel workbook with analysis results.

    Args:
        target_username: Target Instagram username
        followers: List of followers
        following: List of following
        non_mutual: List of non-mutual users

    Returns:
        Styled openpyxl Workbook with 3 sheets:
        1. Non-Mutual (first sheet - user sees immediately)
        2. All Followers
        3. All Following
    """
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    yes_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    no_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Calculate stats
    mutual_count = len(following) - len(non_mutual)
    mutual_percent = (mutual_count / len(following) * 100) if following else 0

    # =====================================================
    # SHEET 1: Non-Mutual (FIRST - user sees this first!)
    # =====================================================
    ws = wb.active
    ws.title = "Не взаимные"

    # Header
    ws.merge_cells("A1:E1")
    ws["A1"] = f"❌ Не взаимные подписки @{target_username}"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    # Stats
    ws["A3"] = "Дата анализа:"
    ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws["A4"] = "Всего подписчиков:"
    ws["B4"] = len(followers)
    ws["A5"] = "Всего подписок:"
    ws["B5"] = len(following)
    ws["A6"] = "Не взаимных:"
    ws["B6"] = len(non_mutual)
    ws["A7"] = "Процент взаимности:"
    ws["B7"] = f"{mutual_percent:.1f}%"
    
    for row in range(3, 8):
        ws[f"A{row}"].font = Font(bold=True)

    ws["A9"] = f"⚠️ Вы подписаны на {len(non_mutual)} аккаунтов, которые НЕ подписаны на вас"
    ws["A9"].font = Font(bold=True, color="C00000")

    # Table headers
    nm_headers = ["#", "Username", "Имя", "Ссылка"]
    for col, header in enumerate(nm_headers, 1):
        cell = ws.cell(row=11, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Data - non-mutual users
    for idx, user in enumerate(non_mutual, 1):
        row = 11 + idx
        ws.cell(row=row, column=1, value=idx).border = border
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=user.username).border = border
        ws.cell(row=row, column=3, value=user.full_name or "").border = border
        
        ig_url = f"https://instagram.com/{user.username}"
        cell = ws.cell(row=row, column=4, value="Открыть")
        cell.hyperlink = ig_url
        cell.font = Font(color="0563C1", underline="single")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 25
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 12

    # =====================================================
    # SHEET 2: All Followers
    # =====================================================
    ws_followers = wb.create_sheet(title="Подписчики")

    # Header
    ws_followers.merge_cells("A1:E1")
    ws_followers["A1"] = f"👥 Все подписчики @{target_username}"
    ws_followers["A1"].font = Font(bold=True, size=16)
    ws_followers["A1"].alignment = Alignment(horizontal="center")

    ws_followers["A3"] = f"Всего подписчиков: {len(followers)}"
    ws_followers["A3"].font = Font(bold=True)

    # Table headers
    followers_headers = ["#", "Username", "Имя", "Вы подписаны?", "Ссылка"]
    for col, header in enumerate(followers_headers, 1):
        cell = ws_followers.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Create following set for lookup
    following_usernames = {f.username.lower() for f in following}

    # Data - followers sorted alphabetically
    sorted_followers = sorted(followers, key=lambda x: x.username.lower())
    for idx, user in enumerate(sorted_followers, 1):
        row = 5 + idx
        is_following_back = user.username.lower() in following_usernames
        
        ws_followers.cell(row=row, column=1, value=idx).border = border
        ws_followers.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws_followers.cell(row=row, column=2, value=user.username).border = border
        ws_followers.cell(row=row, column=3, value=user.full_name or "").border = border
        
        # "Вы подписаны?" column
        follows_text = "✓" if is_following_back else "✗"
        cell = ws_followers.cell(row=row, column=4, value=follows_text)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        cell.fill = yes_fill if is_following_back else no_fill
        
        ig_url = f"https://instagram.com/{user.username}"
        cell = ws_followers.cell(row=row, column=5, value="Открыть")
        cell.hyperlink = ig_url
        cell.font = Font(color="0563C1", underline="single")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    ws_followers.column_dimensions["A"].width = 6
    ws_followers.column_dimensions["B"].width = 25
    ws_followers.column_dimensions["C"].width = 30
    ws_followers.column_dimensions["D"].width = 16
    ws_followers.column_dimensions["E"].width = 12

    # =====================================================
    # SHEET 3: All Following
    # =====================================================
    ws_following = wb.create_sheet(title="Подписки")

    # Header
    ws_following.merge_cells("A1:E1")
    ws_following["A1"] = f"📝 Все подписки @{target_username}"
    ws_following["A1"].font = Font(bold=True, size=16)
    ws_following["A1"].alignment = Alignment(horizontal="center")

    ws_following["A3"] = f"Всего подписок: {len(following)}"
    ws_following["A3"].font = Font(bold=True)

    # Table headers
    following_headers = ["#", "Username", "Имя", "Подписан на вас?", "Ссылка"]
    for col, header in enumerate(following_headers, 1):
        cell = ws_following.cell(row=5, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    # Create followers set for lookup
    follower_usernames = {f.username.lower() for f in followers}

    # Data - following sorted: non-mutual first, then alphabetically
    sorted_following = sorted(
        following, 
        key=lambda x: (x.username.lower() in follower_usernames, x.username.lower())
    )
    
    for idx, user in enumerate(sorted_following, 1):
        row = 5 + idx
        is_follower = user.username.lower() in follower_usernames
        
        ws_following.cell(row=row, column=1, value=idx).border = border
        ws_following.cell(row=row, column=1).alignment = Alignment(horizontal="center")
        ws_following.cell(row=row, column=2, value=user.username).border = border
        ws_following.cell(row=row, column=3, value=user.full_name or "").border = border
        
        # "Подписан на вас?" column
        follower_text = "✓" if is_follower else "✗"
        cell = ws_following.cell(row=row, column=4, value=follower_text)
        cell.border = border
        cell.alignment = Alignment(horizontal="center")
        cell.fill = yes_fill if is_follower else no_fill
        
        ig_url = f"https://instagram.com/{user.username}"
        cell = ws_following.cell(row=row, column=5, value="Открыть")
        cell.hyperlink = ig_url
        cell.font = Font(color="0563C1", underline="single")
        cell.border = border
        cell.alignment = Alignment(horizontal="center")

    ws_following.column_dimensions["A"].width = 6
    ws_following.column_dimensions["B"].width = 25
    ws_following.column_dimensions["C"].width = 30
    ws_following.column_dimensions["D"].width = 16
    ws_following.column_dimensions["E"].width = 12

    return wb


async def generate_xlsx_report(
    check_id: str,
    target_username: str,
    followers: list[InstagramUser],
    following: list[InstagramUser],
    non_mutual: list[InstagramUser],
) -> str:
    """Generate XLSX report file.

    Args:
        check_id: Check UUID
        target_username: Target username
        followers: List of followers
        following: List of following
        non_mutual: List of non-mutual users

    Returns:
        Path to generated file
    """
    # Ensure upload directory exists
    upload_dir = Path(settings.upload_dir)
    upload_dir.mkdir(parents=True, exist_ok=True)

    # Generate filename
    filename = f"{check_id}.xlsx"
    file_path = upload_dir / filename

    # Create workbook
    wb = create_styled_workbook(target_username, followers, following, non_mutual)

    # Save file
    wb.save(file_path)

    logger.info(f"Generated XLSX report: {file_path}")

    return str(file_path)


async def generate_csv_report(
    check_id: str,
    target_username: str,
    non_mutual: list[InstagramUser],
) -> str:
    """Generate simple CSV report.

    Args:
        check_id: Check UUID
        target_username: Target username
        non_mutual: List of non-mutual users

    Returns:
        Path to generated file
    """
    upload_dir = Path(settings.upload_dir)
    upload_dir.mkdir(parents=True, exist_ok=True)

    filename = f"{check_id}.csv"
    file_path = upload_dir / filename

    # Create DataFrame
    data = [
        {
            "#": idx,
            "Username": user.username,
            "Full Name": user.full_name or "",
        }
        for idx, user in enumerate(non_mutual, 1)
    ]

    df = pd.DataFrame(data)
    df.to_csv(file_path, index=False, encoding="utf-8-sig")

    logger.info(f"Generated CSV report: {file_path}")

    return str(file_path)

